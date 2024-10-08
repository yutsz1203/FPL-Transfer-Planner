import math
from functools import partial

import pandas as pd
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from pprint import pprint

#region Constants
COL_A = 1
COL_B = 2
COL_C = 3
COL_D = 4
COL_E = 5
COL_F = 6
COL_G = 7
COL_H = 8
COL_I = 9
COL_J = 10
COL_K = 11
COL_L = 12
COL_M = 13
COL_N = 14
COL_O = 15
COL_P = 16
COL_Q = 17
COL_R = 18
COL_S = 19
COL_T = 20
COL_U = 21
COL_V = 22

NEXT_FIXTURES_COLS = "A:R"
NEXT_GK_COLS = "A:Z"
NEXT_DEF_COLS = "A:AC"
NEXT_MID_COLS = "A:AA"
NEXT_FWD_COLS = "A:AA"

NEXT5_FIXTURES_COLS = "A:H"
NEXT5_GK_COLS = "A:AE"
NEXT5_DEF_COLS = "A:AG"
NEXT5_MID_COLS = "A:AG"
NEXT5_FWD_COLS = "A:AG"

LAST5_GK_COLS = "A:Q"
LAST5_DEF_COLS = "A:U"
LAST5_MID_COLS = "A:U"
LAST5_FWD_COLS = "A:U"

SEASON_TEAMS_COLS = "A:V"
SEASON_GK_COLS = "A:AI"
SEASON_DEF_COLS = "A:AS"
SEASON_MID_COLS = "A:AS"
SEASON_FWD_COLS = "A:AS"

whiteFill = PatternFill(start_color='FFFFFFFF',end_color='FFFFFFFF',fill_type='solid')
redFill = PatternFill(start_color='FFFF0000',end_color='FFFF0000', fill_type='solid')
yellowFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00',fill_type='solid')
greenFill = PatternFill(start_color='ff00b050',end_color='ff00b050',fill_type='solid')
lightGreenFill = PatternFill(start_color='ffc6e0b4',end_color='ffc6e0b4',fill_type='solid')
blueFill = PatternFill(start_color='ff2f75b5',end_color='ff2f75b5',fill_type='solid')
lightBlueFill = PatternFill(start_color='ffbdd7ee',end_color='ffbdd7ee',fill_type='solid')

DEFAULT_FONT = Font(name='Calibri', size=16)

BASE_URL = "https://fantasy.premierleague.com/api/bootstrap-static/"
ID = 1185677
CURRENT_EXCEL = "Reports/GW Report.xlsx"
NEXT_EXCEL = "Reports/GW Preview.xlsx"
LAST5_EXCEL = "Reports/Last5 Report.xlsx"
NEXT5_EXCEL = "Reports/Next5 Preview.xlsx"
SEASON_EXCEL = "Reports/Season Report.xlsx"
MINUTES_THRESHOLD = 30
GK_COUNT = 80
DEF_COUNT = 250
MID_COUNT = 350
FWD_COUNT = 100

team_map = {
        "ARS": 2,
        "AVL": 3,
        "BOU": 4,
        "BRE": 5,
        "BHA": 6,
        "CHE": 7,
        "CRY": 8,
        "EVE": 9,
        "FUL": 10,
        "IPS": 11,
        "LEI": 12,
        "LIV": 13,
        "MCI": 14,
        "MUN": 15,
        "NEW": 16,
        "NFO": 17,
        "SOU": 18,
        "TOT": 19,
        "WHU": 20,
        "WOL": 21
        }

teams = {
        1: "Arsenal",
        2: "Aston Villa",
        3: "Bournemouth",
        4: "Brentford",
        5: "Brighton",
        6: "Chelsea",
        7: "Crystal Palace",
        8: "Everton",
        9: "Fulham",
        10: "Ipswich",
        11: "Leicester",
        12: "Liverpool",
        13: "Man City",
        14: "Man Utd",
        15: "Newcastle",
        16: "Nott'm Forest",
        17: "Southampton",
        18: "Spurs",
        19: "West Ham",
        20: "Wolves"
    }

team_full_from = {
    "ARS": "Arsenal",
    "AVL": "Aston Villa",
    "BOU": "Bournemouth",
    "BRE": "Brentford",
    "BHA": "Brighton",
    "CHE": "Chelsea",
    "CRY": "Crystal Palace",
    "EVE": "Everton",
    "FUL": "Fulham",
    "IPS": "Ipswich",
    "LEI": "Leicester",
    "LIV": "Liverpool",
    "MCI": "Man City",
    "MUN": "Man Utd",
    "NEW": "Newcastle",
    "NFO": "Nott'm Forest",
    "SOU": "Southampton",
    "TOT": "Spurs",
    "WHU": "West Ham",
    "WOL": "Wolves"
}
position = {
        1: "GK",
        2: "DEF",
        3: "MID",
        4: "FWD"
    }

team_oi = {}
team_di = {}
team_h_a_count = {}

#region Helper functions
def get_fmt_str(x, fill):
        return "{message: >{fill}}".format(message=x, fill=fill)

def format_columns(df):
    s = df.astype(str).agg(lambda x: x.str.len()).max() 

    pad = 6  
    fmts = {}
    for idx, c_len in s.items():
        if isinstance(idx, tuple):
            lab_len = max([len(str(x)) for x in idx])
        else:
            lab_len = len(str(idx))

        fill = max(lab_len, c_len) + pad - 1
        fmts[idx] = partial(get_fmt_str, fill=fill)
    return fmts

def write_df_to_sheet(df, sheet):
    for col, value in enumerate(df.columns, start=1):
        cell = sheet.cell(row=1, column=col, value=value)
        cell.font = DEFAULT_FONT
    for row, value in enumerate(df.values, start=2):
        for col, cell in enumerate(value, start=1):
            cell = sheet.cell(row=row, column=col, value=cell)
            cell.font = DEFAULT_FONT

#region Get data functions
def get_gameweek():
    gws = requests.get(BASE_URL).json()["events"]
    for gw in gws:
        if gw["is_current"]:
            return gw["id"]

def get_players():
    r = requests.get(BASE_URL).json()
    players = r["elements"] # 1 to 642
    gks, defs, mids, fwds = [], [], [], []
    for player in players:
        id = player["id"]
        individual = requests.get(f"https://fantasy.premierleague.com/api/element-summary/{id}/").json()["history"]
        if len(individual) == 1 and individual[0]["round"] == get_gameweek():
            if player["element_type"] == 1:
                gks.append([player["web_name"], teams[player["team"]],
                (player["now_cost"] / 10.0),
                position[player["element_type"]]])
            elif player["element_type"] == 2:
                defs.append([player["web_name"], teams[player["team"]],
                (player["now_cost"] / 10.0),
                position[player["element_type"]]])
            elif player["element_type"] == 3:
                mids.append([player["web_name"], teams[player["team"]],
                (player["now_cost"] / 10.0),
                position[player["element_type"]]])
            if player["element_type"] == 4:
                fwds.append([player["web_name"], teams[player["team"]],
                (player["now_cost"] / 10.0),
                position[player["element_type"]]])
            print(player["web_name"])
    gk_df = pd.DataFrame(gks, columns=["Name", "Team", "Price", "Pos"])
    def_df = pd.DataFrame(defs, columns=["Name", "Team", "Price", "Pos"])
    mid_df = pd.DataFrame(mids, columns=["Name", "Team", "Price", "Pos"])
    fwd_df = pd.DataFrame(fwds, columns=["Name", "Team", "Price", "Pos"])

    gk_df.to_excel("NewPlayers/GK.xlsx", sheet_name="GK List", index=False)
    def_df.to_excel("NewPlayers/DEF.xlsx", sheet_name="DEF List", index=False)
    mid_df.to_excel("NewPlayers/MID.xlsx", sheet_name="MID List", index=False)
    fwd_df.to_excel("NewPlayers/FWD.xlsx", sheet_name="FWD List", index=False)

#region Main functions
def update_results_current():
    url = "https://fpl.page/bonus"
    page = requests.get(url)

    soup = BeautifulSoup(page.text, "html.parser")
    
    fixtures = soup.find_all("li", class_="fixture-item")

    df = pd.DataFrame(columns=["Home", "H_G", "A_G", "Away"])

    wb = load_workbook(CURRENT_EXCEL)
    sheet = wb["Results"]

    for row, fixture in enumerate(fixtures, start=2):
        home_team = fixture.find("span", class_="home-text").text.strip()
        away_team = fixture.find("span", class_="away-text").text.strip()
        score_box = fixture.find("span", class_="score-box")

        home_returners = fixture.find_all("div", class_="scorers-box-home")
        away_returners = fixture.find_all("div", class_="scorers-box-away")

        home_scorers, home_assists, away_scorers, away_assists, bonus_list = "", "", "", "", ""

        if home_returners:
            for returner in home_returners:
                scorers = returner.find_all("li", class_="scorer-item-goals")
                passers = returner.find_all("li", class_="scorer-item")
                if scorers:
                    for scorer in scorers:
                        goals = len(scorer.find_all("svg", attrs={"data-icon": "futbol"}))
                        if goals > 0:
                            home_scorers += f" {scorer.text.strip()} ({goals})"
                if passers:
                        for passer in passers:
                            assists = len(passer.find_all("svg", attrs={"data-icon": "a"}))
                            if assists > 0:
                                home_assists += f" {passer.text.strip()} ({assists})"
        bonuses = fixture.find_all("li", class_="bonus-item")
        for bonus in bonuses:
            bonus_list += f" {bonus.text.strip()}"
                
        if away_returners:
            for returner in away_returners:
                individuals = returner.find_all("li", class_="scorer-item")
                if individuals:
                    for individual in individuals:
                        goals = len(individual.find_all("svg", attrs={"data-icon": "futbol"}))
                        assists = len(individual.find_all("svg", attrs={"data-icon": "a"}))
                        
                        if goals > 0:
                            away_scorers += f" {individual.text.strip()} ({goals})"
                        
                        if assists > 0:
                            away_assists += f" {individual.text.strip()} ({assists})"
        
        if score_box:
            scores = score_box.text.split("-")
            scores = list(map(int, scores))
            
            df = df._append({"Home": home_team, "H_G": scores[0], "A_G": scores[1], "Away": away_team}, ignore_index=True)
            print(f"{home_team} {scores[0]} - {scores[1]} {away_team}")

        sheet[f"E{row}"] = home_scorers 
        sheet[f"F{row}"] = home_assists
        sheet[f"G{row}"] = away_scorers
        sheet[f"H{row}"] = away_assists
        sheet[f"I{row}"] = bonus_list
        
    write_df_to_sheet(df, sheet)

    print("Finished updating current GW results.")
    wb.save(CURRENT_EXCEL)

def update_fixture_next():
    #web scrapping
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get("https://fpl.page/fixtureTicker")

    #get fixtures
    table = driver.find_element(By.CSS_SELECTOR, ".transfers-body") 
    table_soup = BeautifulSoup(table.get_attribute("innerHTML"), "html.parser")
    teams = table_soup.find_all("tr")

    fixtures = []

    for team in teams:
        team_name = team.find("td", class_="fdr-teamNames").text
        td = team.find_all("td")[1]
        opponent = td.find("span").text if td.find("span") else None
        if opponent:
            opponent_name = opponent.split(" ")[0]
            side = opponent.split(" ")[1]
        fixtures.append({team_name: opponent_name, "side": side})
    print(fixtures)
    print("Finished getting fixtures.")
    driver.quit()

    fixture_df = pd.read_excel(NEXT_EXCEL, sheet_name="Fixtures", usecols="A:O", header=0, nrows=20)
    team_df = pd.read_excel(SEASON_EXCEL, sheet_name="Teams", usecols="A:V", header=0, nrows=20)
    team_sheet = load_workbook(SEASON_EXCEL)["Teams"]
    gk_df = pd.read_excel(SEASON_EXCEL, sheet_name="GK", usecols=SEASON_GK_COLS, header=0, nrows=GK_COUNT)
    def_df = pd.read_excel(SEASON_EXCEL, sheet_name="DEF", usecols=SEASON_DEF_COLS, header=0, nrows=DEF_COUNT)
    mid_df = pd.read_excel(SEASON_EXCEL, sheet_name="MID", usecols=SEASON_MID_COLS, header=0, nrows=MID_COUNT)
    fwd_df = pd.read_excel(SEASON_EXCEL, sheet_name="FWD", usecols=SEASON_FWD_COLS, header=0, nrows=FWD_COUNT)

    columns_to_update = ["H_Oi", "H_Di", "A_Oi", "A_Di", "Oi", "Di"]
    opponent_columns_to_update = ["O_H_Oi", "O_H_Di", "O_A_Oi", "O_A_Di", "O_Oi", "O_Di"]

    match_up_rows = [[1,2], [18,19], [35,36], [52,53], [69,70]]
    i = 0
    match_up_cols = [["T", "U", "V"], ["AB", "AC", "AD"]]
    j = 0

    wb = load_workbook(NEXT_EXCEL)
    fixture_sheet = wb["Fixtures"]

    for fixture in fixtures:
        team_name = list(fixture.keys())[0]
        opponent = team_full_from[fixture[team_name]]
        side = fixture["side"]
        team_row = team_df.index[team_df["Team"] == team_name].tolist()[0]
        fixture_team_row = fixture_df.index[fixture_df["Team"] == team_name].tolist()[0]

        best_def = def_df[def_df["Team"]==team_name].sort_values(by=["xGI"], ascending=False).head(1)["Name"].values[0]
        best_mid = mid_df[mid_df["Team"]==team_name].sort_values(by=["xGI"], ascending=False).head(1)["Name"].values[0]
        best_fwd = fwd_df[fwd_df["Team"]==team_name].sort_values(by=["xGI"], ascending=False).head(1)["Name"].values[0]

        opponent_team_row = team_df.index[team_df["Team"] == opponent].tolist()[0]
        fixture_df.loc[fixture_team_row, "Opponent"] = opponent

        for column, opp_column in zip(columns_to_update, opponent_columns_to_update):
            fixture_df.loc[fixture_team_row, "Side"] = side
            fixture_df.loc[fixture_team_row, "DEF"] = best_def
            fixture_df.loc[fixture_team_row, "MID"] = best_mid
            fixture_df.loc[fixture_team_row, "FWD"] = best_fwd
            fixture_df.loc[fixture_team_row, column] = team_df.loc[team_row, column]
            fixture_df.loc[fixture_team_row, opp_column] = team_df.loc[opponent_team_row, column]
        
        if side == "(H)":
            if i == 5:
                i = 0
                j = 1
            fixture_sheet[f"{match_up_cols[j][0]}{match_up_rows[i][0]}"] = team_name
            fixture_sheet[f"{match_up_cols[j][1]}{match_up_rows[i][0]}"] = fixture_df.loc[fixture_team_row, "Oi"]
            fixture_sheet[f"{match_up_cols[j][2]}{match_up_rows[i][0]}"] = fixture_df.loc[fixture_team_row, "Di"]

            fixture_sheet[f"{match_up_cols[j][0]}{match_up_rows[i][1]}"] = opponent
            fixture_sheet[f"{match_up_cols[j][1]}{match_up_rows[i][1]}"] = fixture_df.loc[fixture_team_row, "O_Oi"]
            fixture_sheet[f"{match_up_cols[j][2]}{match_up_rows[i][1]}"] = fixture_df.loc[fixture_team_row, "O_Di"]
            i += 1

    fixture_df = fixture_df.sort_values(by=["O_Di"], ascending=False)
    print(fixture_df)
 
    lod = team_sheet["H24"].value
    fixture_sheet["B23"] = lod
    write_df_to_sheet(fixture_df, fixture_sheet)
    wb.save(NEXT_EXCEL)
    print("Finished updating next fixture data.")


def update_player_next():
    fixture_df = pd.read_excel(NEXT_EXCEL, sheet_name="Fixtures", usecols="A:R", header=0, nrows=20)
    lod = load_workbook(SEASON_EXCEL, data_only=True)["Teams"]["H24"].value
    
    gk_df = pd.read_excel(SEASON_EXCEL, sheet_name="GK", usecols=SEASON_GK_COLS, header=0, nrows=GK_COUNT)
    gk_df.drop(["Starts","H_Points", "A_Points", "H_CS", "A_CS", "H_Goals Conceded", "A_Goals Conceded", "H_xG Conceded", "A_xG Conceded", "H_Saves", "A_Saves", "H_Bonus", "A_Bonus"], axis=1, inplace=True)
    # Add home / away projected goals here after gw10
    gk_df["Opponent"] = gk_df["Team"].apply(lambda team: fixture_df.loc[fixture_df["Team"] == team, "Opponent"].values[0])
    gk_df["O_Oi"] = gk_df["Team"].apply(lambda team: fixture_df.loc[fixture_df["Team"] == team, "O_Oi"].values[0])
    gk_df["Di"] = gk_df["Team"].apply(lambda team: fixture_df.loc[fixture_df["Team"] == team, "Di"].values[0])
    gk_df["Projected Performance"] = -(gk_df["O_Oi"] * gk_df["Di"] * lod - (gk_df["xG Prevented"] / get_gameweek()))  
    gk_df.sort_values(by=["Projected Performance"], ascending=False, inplace=True)
    gk_df = gk_df.drop(gk_df[gk_df["Games"]==0].index)

    outfield_cols = ["H_Points", "A_Points", "H_G", "A_G", "H_xG", "A_xG", "H_xGDiff", "A_xGDiff", "H_A", "A_A", "H_xA", "A_xA", "H_xADiff", "A_xADiff", "H_xGIDiff", "A_xGIDiff", "xGDiff", "xADiff", "H_Bonus", "A_Bonus", "Bonus"]
    def_df = pd.read_excel(SEASON_EXCEL, sheet_name="DEF", usecols=SEASON_DEF_COLS, header=0, nrows=DEF_COUNT)
    mid_df = pd.read_excel(SEASON_EXCEL, sheet_name="MID", usecols=SEASON_MID_COLS, header=0, nrows=MID_COUNT)
    fwd_df = pd.read_excel(SEASON_EXCEL, sheet_name="FWD", usecols=SEASON_FWD_COLS, header=0, nrows=FWD_COUNT)
    
    def make_df(df, defender=False):
        df.drop(outfield_cols, axis=1, inplace=True)
        df["Opponent"] = df["Team"].apply(lambda team: fixture_df.loc[fixture_df["Team"] == team, "Opponent"].values[0])
        if defender:
            df["O_Oi"] = df["Team"].apply(lambda team: fixture_df.loc[fixture_df["Team"] == team, "O_Oi"].values[0])
            df["O_Di"] = df["Team"].apply(lambda team: fixture_df.loc[fixture_df["Team"] == team, "O_Di"].values[0])
            df["Di"] = df["Team"].apply(lambda team: fixture_df.loc[fixture_df["Team"] == team, "Di"].values[0])
            df["Projected Performance"] = ((df["xGI"] / get_gameweek())) * df["O_Di"] * lod - (df["O_Oi"] * df["Di"] * lod)
        else:
            df["O_Di"] = df["Team"].apply(lambda team: fixture_df.loc[fixture_df["Team"] == team, "O_Di"].values[0])
            df["Projected Performance"] = ((df["xGI"] / get_gameweek())) * df["O_Di"] * lod 
        
        df.sort_values(by=["Projected Performance"], ascending=False, inplace=True)

    make_df(def_df, True)
    def_df = def_df.drop(def_df[def_df["Games"]==0].index)
    make_df(mid_df)
    mid_df = mid_df.drop(mid_df[mid_df["Games"]==0].index)
    make_df(fwd_df)
    fwd_df = fwd_df.drop(fwd_df[fwd_df["Games"]==0].index)

    wb = load_workbook(NEXT_EXCEL)
    write_df_to_sheet(gk_df, wb["GK"])
    print("Finished updating next GW GK data.")
    write_df_to_sheet(def_df, wb["DEF"])
    print("Finished updating next GW DEF data.")
    write_df_to_sheet(mid_df, wb["MID"])
    print("Finished updating next GW MID data.")
    write_df_to_sheet(fwd_df, wb["FWD"])
    print("Finished updating next GW FWD data.")
    wb.save(NEXT_EXCEL)


def update_my_team_next():
    gw = get_gameweek()
    url = f"https://fantasy.premierleague.com/api/entry/{ID}/event/{gw}/picks/"
    team = requests.get(url).json()["picks"]
    player_ids = [player["element"] for player in team]
    r = requests.get(BASE_URL).json()["elements"]
    players = [player for player in r if player["id"] in player_ids]

    wb = load_workbook(NEXT_EXCEL)
    gk_df = pd.read_excel(NEXT_EXCEL, sheet_name="GK", usecols="A:Z", header=0, nrows=GK_COUNT)
    def_df = pd.read_excel(NEXT_EXCEL, sheet_name="DEF", usecols="A:AC", header=0, nrows=DEF_COUNT)
    mid_df = pd.read_excel(NEXT_EXCEL, sheet_name="MID", usecols="A:AA", header=0, nrows=MID_COUNT)
    fwd_df = pd.read_excel(NEXT_EXCEL, sheet_name="FWD", usecols="A:AA", header=0, nrows=FWD_COUNT)

    team_gk_rows = []
    team_def_rows = []
    team_mid_rows = []
    team_fwd_rows = []

    for player in players:
        if player["element_type"] == 1:
            team_gk_rows.append(gk_df.loc[gk_df["Name"] == player["web_name"]])
        elif player["element_type"] == 2:
            team_def_rows.append(def_df.loc[def_df["Name"] == player["web_name"]])
        elif player["element_type"] == 3:
            team_mid_rows.append(mid_df.loc[mid_df["Name"] == player["web_name"]])
        elif player["element_type"] == 4:
            team_fwd_rows.append(fwd_df.loc[fwd_df["Name"] == player["web_name"]])
        print("Finished getting " + player["web_name"])

    team_gk_df = pd.concat(team_gk_rows, ignore_index=True)
    team_def_df = pd.concat(team_def_rows, ignore_index=True)
    team_mid_df = pd.concat(team_mid_rows, ignore_index=True)
    team_fwd_df = pd.concat(team_fwd_rows, ignore_index=True)

    team_df = pd.concat([team_def_df, team_mid_df, team_fwd_df, team_gk_df], ignore_index=True, sort=False)
    team_df = team_df.sort_values(by=["Projected Performance"], ascending=False)

    team_season = wb["My Team"]
    write_df_to_sheet(team_df, team_season)
    wb.save(NEXT_EXCEL)
    print("Finished updating my team - next GW")



def update_fixture():
    #web scrapping
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get("https://fpl.page/fixtureTicker")

    #get gameweeks
    headers = driver.find_elements(By.CSS_SELECTOR, "th.fdr-teamNames")
    combined_headers = "".join([header.get_attribute("innerHTML") for header in headers[0:5]])
    header_soup = BeautifulSoup(combined_headers, "html.parser")
    gameweeks = ["GW" + text.strip("GW") for text in header_soup.stripped_strings if text != "GW"]
    print("Finished getting gameweeks.")

    #get fixtures
    table = driver.find_element(By.CSS_SELECTOR, ".transfers-body") 
    table_soup = BeautifulSoup(table.get_attribute("innerHTML"), "html.parser")
    teams = table_soup.find_all("tr")

    fixtures = []

    for team in teams:
        team_name = team.find("td", class_="fdr-teamNames").text
        tds = team.find_all("td")[1:6]
        opponents = [td.find("span").text for td in tds if td.find("span")]
        fixtures.append({"Team": team_name, "Next 5": opponents})
    print("Finished getting fixtures.")

    driver.quit()

    #update excel
    wb = load_workbook(NEXT5_EXCEL, data_only=True)
    sheet = wb["Fixtures"]

    for row in sheet.iter_rows(min_row=2, max_row=21, min_col=1, max_col=8):
        for cell in row:
            cell.fill = whiteFill

    for col, gameweek in enumerate(gameweeks, start=2):
        sheet.cell(row=1, column=col, value=gameweek)

    season_wb = load_workbook(SEASON_EXCEL)
    team_data_sheet = season_wb["Teams"]
    teams_df = pd.read_excel(SEASON_EXCEL, sheet_name="Teams", usecols="A:V", header=0, nrows=20)
    for row, team in enumerate(fixtures, start=2):
        sheet[f"A{row}"] = team["Team"]
        sheet[f"G{row}"] = 0
        sheet[f"H{row}"] = 0
        sheet[f"I{row}"] = 0
        sheet[f"J{row}"] = 0
        # home_count = 0
        # away_count = 0
        for col, opponent in enumerate(team["Next 5"], start=2):
            #filling in opponents
            sheet.cell(row=row, column=col, value=opponent)

            #aggregating Oi and Di
            cleaned_opponent = opponent.split(" ")[0]
            opponent_team = team_full_from[cleaned_opponent]
            #gets opponent's Away Oi, Di
            #use after gw18
            # if(opponent.split(" ")[1] == "(H)"):
            #     home_count += 1
            #     sheet[f"G{row}"] = sheet[f"G{row}"].value + team_data_sheet[f"S{opponent_team_row}"].value
            #     sheet[f"H{row}"] = sheet[f"H{row}"].value + team_data_sheet[f"T{opponent_team_row}"].value
            # else:
            #     away_count += 1
            #     sheet[f"G{row}"] = sheet[f"G{row}"].value + team_data_sheet[f"Q{opponent_team_row}"].value
            #     sheet[f"H{row}"] = sheet[f"H{row}"].value + team_data_sheet[f"R{opponent_team_row}"].value
            # team_h_a_count[fixtures[row - 2]["Team"]] = [home_count, away_count]
            sheet[f"G{row}"] = sheet[f"G{row}"].value + teams_df.loc[teams_df["Team"] == opponent_team, "Oi"].values[0]
            sheet[f"H{row}"] = sheet[f"H{row}"].value + teams_df.loc[teams_df["Team"] == opponent_team, "Di"].values[0]
        sheet[f"I{row}"] = sheet[f"H{row}"].value * teams_df.loc[teams_df["Team"] == team["Team"], "Oi"].values[0] * 5
        sheet[f"J{row}"] = sheet[f"G{row}"].value * teams_df.loc[teams_df["Team"] == team["Team"], "Di"].values[0] * 5
        print(f"Finished updating {team['Team']}")
    
    wb.save(NEXT5_EXCEL)
    df = pd.read_excel(NEXT5_EXCEL, sheet_name="Fixtures", usecols="G:J", header=0, nrows=20)

    sheet["G23"].value = df["Oi"].mean()
    sheet["H23"].value = df["Di"].mean()
    sheet["I23"].value = df["Expected Goals"].mean()
    sheet["J23"].value = df["Expected Goals Conceded"].mean()

    low_oi_index = df.sort_values(by=["Oi"], ascending=True)["Oi"].head(5).index.tolist()
    high_oi_index = df.sort_values(by=["Oi"], ascending=False)["Oi"].head(5).index.tolist()
    high_di_index = df.sort_values(by=["Di"], ascending=False)["Di"].head(5).index.tolist()
    low_di_index = df.sort_values(by=["Di"], ascending=True)["Di"].head(5).index.tolist()

    for col, (low_oi_idx, high_di_idx) in enumerate(zip(low_oi_index, high_di_index), start=2):
        sheet.cell(row=24, column=col, value=sheet[f"A{high_di_idx + 2}"].value)
        sheet.cell(row=25, column=col, value=sheet[f"A{low_oi_idx + 2}"].value)
    
    for high_oi_idx, low_oi_idx, high_di_idx, low_di_idx in zip(high_oi_index, low_oi_index, high_di_index, low_di_index):
        sheet.cell(row=high_oi_idx + 2, column=COL_G).fill = redFill
        sheet.cell(row=low_di_idx + 2, column=COL_H).fill = redFill

        sheet.cell(row=low_oi_idx + 2, column=COL_G).fill = greenFill
        sheet.cell(row=low_oi_idx + 2, column=COL_A).fill = greenFill

        sheet.cell(row=high_di_idx + 2, column=COL_H).fill = blueFill
        sheet.cell(row=high_di_idx + 2, column=COL_A).fill = blueFill

        if low_oi_idx in high_di_index:
            sheet.cell(row=low_oi_idx + 2, column=COL_A).fill = yellowFill
        
        if high_di_idx in low_oi_index:
            sheet.cell(row=high_di_idx + 2, column=COL_A).fill = yellowFill
        
        if high_oi_idx in low_di_index:
            sheet.cell(row=high_oi_idx + 2, column=COL_A).fill = redFill
        
        if low_di_idx in high_oi_index:
            sheet.cell(row=low_di_idx + 2, column=COL_A).fill = redFill

    print("Finished updating fixture data.")
    wb.save(NEXT5_EXCEL)


def update_player_next5():
    #web scrapping
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get("https://fpl.page/fixtureTicker")

    #get gameweeks
    headers = driver.find_elements(By.CSS_SELECTOR, "th.fdr-teamNames")
    combined_headers = "".join([header.get_attribute("innerHTML") for header in headers[0:5]])
    header_soup = BeautifulSoup(combined_headers, "html.parser")
    gameweeks = ["GW" + text.strip("GW") for text in header_soup.stripped_strings if text != "GW"]

    #get fixtures
    table = driver.find_element(By.CSS_SELECTOR, ".transfers-body") 
    table_soup = BeautifulSoup(table.get_attribute("innerHTML"), "html.parser")
    teams = table_soup.find_all("tr")

    fixtures = {}

    for team in teams:
        team_name = team.find("td", class_="fdr-teamNames").text
        tds = team.find_all("td")[1:6]
        opponents = [td.find("span").text for td in tds if td.find("span")]
        fixtures[team_name] = opponents

    driver.quit()

    for team in fixtures:
        home_count = 0
        away_count = 0
        for opponent in fixtures[team]:
            #gets opponent's Away Oi, Di
            if(opponent.split(" ")[1] == "(H)"):
                home_count += 1
            else:
                away_count += 1
            
        team_h_a_count[team] = [home_count, away_count]

    fixture_df = pd.read_excel(NEXT5_EXCEL, sheet_name="Fixtures", usecols="A:H", header=0, nrows=20)
    for team in fixture_df["Team"].values.tolist():
         team_oi[team] = fixture_df.loc[fixture_df["Team"] == team, "Oi"].values[0]
         team_di[team] = fixture_df.loc[fixture_df["Team"] == team, "Di"].values[0]

    lod = load_workbook(SEASON_EXCEL, data_only=True)["Teams"]["H24"].value
    teams_df = pd.read_excel(SEASON_EXCEL, sheet_name="Teams", usecols=SEASON_TEAMS_COLS, header=0, nrows=20)

    gk_df = pd.read_excel(SEASON_EXCEL, sheet_name="GK", usecols=SEASON_GK_COLS, header=0, nrows=GK_COUNT)
    gk_df.drop(["Starts","H_Points", "A_Points", "H_CS", "A_CS", "H_Goals Conceded", "A_Goals Conceded", "H_xG Conceded", "A_xG Conceded", "H_Saves", "A_Saves", "H_Bonus", "A_Bonus"], axis=1, inplace=True)
    gk_df["Oi"] = gk_df["Team"].apply(lambda team: team_oi.get(team, ""))  
    gk_df["Di"] = gk_df["Team"].apply(lambda team: teams_df.loc[teams_df["Team"] == team, "Di"].values[0])
    gk_df["Di"] *= 5

    h_xg_prevented_per_game = gk_df["H_xG Prevented"] / gk_df["H_Games"]
    a_xg_prevented_per_game = gk_df["A_xG Prevented"] / gk_df["A_Games"]
    h_xg_prevented_per_game.replace([np.inf, -np.inf, np.nan], 0, inplace=True)
    a_xg_prevented_per_game.replace([np.inf, -np.inf, np.nan], 0, inplace=True)

    gk_df["next5_h"] = gk_df["Team"].apply(lambda team: team_h_a_count.get(team, [0,0])[0])
    gk_df["next5_a"] = gk_df["Team"].apply(lambda team: team_h_a_count.get(team, [0,0])[1])
    gk_df["Projected Goals"] = gk_df["Oi"] * gk_df["Di"] * lod - ((gk_df["xG Prevented"]/get_gameweek()) * 5)
    gk_df["H/A Weighted Performance"] = gk_df["Oi"] * gk_df["Di"] * lod - (h_xg_prevented_per_game * gk_df["next5_h"] + a_xg_prevented_per_game * gk_df["next5_a"])
    gk_df.drop(["next5_h", "next5_a"], axis=1, inplace=True)
    gk_df.sort_values(by=["Projected Goals"], ascending=True, inplace=True)
    gk_df = gk_df.drop(gk_df[gk_df["Games"]==0].index)

    for i, gameweek in enumerate(gameweeks):
        gk_df[gameweek] = gk_df["Team"].apply(lambda team: fixtures.get(team, [])[i] if team in fixtures else None)

    outfield_cols = ["H_Points", "A_Points", "H_G", "A_G", "H_xG", "A_xG", "H_xGDiff", "A_xGDiff", "H_A", "A_A", "H_xA", "A_xA", "H_xADiff", "A_xADiff", "H_xGIDiff", "A_xGIDiff", "xGDiff", "xADiff", "H_Bonus", "A_Bonus", "Bonus"]
    def_df = pd.read_excel(SEASON_EXCEL, sheet_name="DEF", usecols=SEASON_DEF_COLS, header=0, nrows=DEF_COUNT)
    mid_df = pd.read_excel(SEASON_EXCEL, sheet_name="MID", usecols=SEASON_MID_COLS, header=0, nrows=MID_COUNT)
    fwd_df = pd.read_excel(SEASON_EXCEL, sheet_name="FWD", usecols=SEASON_FWD_COLS, header=0, nrows=FWD_COUNT)
    
    #Change calculation method for defenders, copy from update_player_next.
    def make_df(df):
        df.drop(outfield_cols, axis=1, inplace=True)
        df["Oi"] = df["Team"].apply(lambda team: teams_df.loc[teams_df["Team"] == team, "Oi"].values[0])
        df["Oi"] *= 5
        df["Di"] = df["Team"].apply(lambda team: team_di.get(team, ""))

        h_xgi_per_game = df["H_xGI"] / df["H_Games"]
        a_xgi_per_game = df["A_xGI"] / df["A_Games"]
        h_xgi_per_game.replace([np.inf, -np.inf, np.nan], 0, inplace=True)
        a_xgi_per_game.replace([np.inf, -np.inf, np.nan], 0, inplace=True)

        df["next5_h"] = df["Team"].apply(lambda team: team_h_a_count.get(team, [0,0])[0])
        df["next5_a"] = df["Team"].apply(lambda team: team_h_a_count.get(team, [0,0])[1])
        df["Projected Performance"] = ((df["xGI"] / get_gameweek()) * 5) * df["Di"] * lod 
        df["H/A Weighted Performance"] = df["Di"] * lod * (h_xgi_per_game * df["next5_h"] + a_xgi_per_game * df["next5_a"])
        df.drop(["next5_h", "next5_a"], axis=1, inplace=True)
        df.sort_values(by=["Projected Performance"], ascending=False, inplace=True)

        for i, gameweek in enumerate(gameweeks):
            df[gameweek] = df["Team"].apply(lambda team: fixtures.get(team, [])[i] if team in fixtures else None)

    
    make_df(def_df)
    def_df = def_df.drop(def_df[def_df["Games"]==0].index)
    make_df(mid_df)
    mid_df = mid_df.drop(mid_df[mid_df["Games"]==0].index)
    make_df(fwd_df)
    fwd_df = fwd_df.drop(fwd_df[fwd_df["Games"]==0].index)

    wb = load_workbook(NEXT5_EXCEL)
    write_df_to_sheet(gk_df, wb["GK"])
    print("Finished updating Next 5 GK data.")
    write_df_to_sheet(def_df, wb["DEF"])
    print("Finished updating Next 5 DEF data.")
    write_df_to_sheet(mid_df, wb["MID"])
    print("Finished updating Next 5 MID data.")
    write_df_to_sheet(fwd_df, wb["FWD"])
    print("Finished updating Next 5 FWD data.")
    wb.save(NEXT5_EXCEL)


def update_my_team_next5():
    gw = get_gameweek()
    url = f"https://fantasy.premierleague.com/api/entry/{ID}/event/{gw}/picks/"
    team = requests.get(url).json()["picks"]
    player_ids = [player["element"] for player in team]
    r = requests.get(BASE_URL).json()["elements"]
    players = [player for player in r if player["id"] in player_ids]

    wb = load_workbook(NEXT5_EXCEL)
    gk_df = pd.read_excel(NEXT5_EXCEL, sheet_name="GK", usecols=NEXT5_GK_COLS, header=0, nrows=GK_COUNT)
    def_df = pd.read_excel(NEXT5_EXCEL, sheet_name="DEF", usecols=NEXT5_DEF_COLS, header=0, nrows=DEF_COUNT)
    mid_df = pd.read_excel(NEXT5_EXCEL, sheet_name="MID", usecols=NEXT5_MID_COLS, header=0, nrows=MID_COUNT)
    fwd_df = pd.read_excel(NEXT5_EXCEL, sheet_name="FWD", usecols=NEXT5_FWD_COLS, header=0, nrows=FWD_COUNT)

    team_gk_rows = []
    team_def_rows = []
    team_mid_rows = []
    team_fwd_rows = []

    for player in players:
        if player["element_type"] == 1:
            team_gk_rows.append(gk_df.loc[gk_df["Name"] == player["web_name"]])
        elif player["element_type"] == 2:
            team_def_rows.append(def_df.loc[def_df["Name"] == player["web_name"]])
        elif player["element_type"] == 3:
            team_mid_rows.append(mid_df.loc[mid_df["Name"] == player["web_name"]])
        elif player["element_type"] == 4:
            team_fwd_rows.append(fwd_df.loc[fwd_df["Name"] == player["web_name"]])
        print("Finished getting " + player["web_name"])

    team_gk_df = pd.concat(team_gk_rows, ignore_index=True)
    team_def_df = pd.concat(team_def_rows, ignore_index=True)
    team_mid_df = pd.concat(team_mid_rows, ignore_index=True)
    team_fwd_df = pd.concat(team_fwd_rows, ignore_index=True)

    team_df = pd.concat([team_def_df, team_mid_df, team_fwd_df, team_gk_df], ignore_index=True, sort=False)
    team_season = wb["My Team"]

    write_df_to_sheet(team_df, team_season)
    wb.save(NEXT5_EXCEL)
    print("Finished updating my team - next 5")


def show_summary():
    wb = load_workbook(SEASON_EXCEL)
    team_data_sheet = wb["Team Data"]
    fixture_sheet = wb["Fixtures"]

    team_df = pd.read_excel(SEASON_EXCEL, sheet_name="Team Data", usecols=SEASON_TEAMS_COLS, header=0, nrows=20, dtype={"Team": str, "GF": int, "GA": int, "Games": int, "Avg.GF": float, "Avg.GA": float, "Oi": float, "Di": float})
    top_attacking_team = []
    top_defending_team = []

    columns = ["B", "C", "D", "E", "F"]

    for col in columns:
        attack_team = team_data_sheet[f"{col}26"].value
        defend_team = team_data_sheet[f"{col}27"].value 
        top_attacking_team.append({"Team": attack_team, "Oi": team_df.loc[team_df["Team"] == attack_team, "Oi"].values[0], "Di": team_df.loc[team_df["Team"] == attack_team, "Di"].values[0]})
        top_defending_team.append({"Team": defend_team, "Oi": team_df.loc[team_df["Team"] == defend_team, "Oi"].values[0], "Di": team_df.loc[team_df["Team"] == defend_team, "Di"].values[0]})

    fixture_df = pd.read_excel(SEASON_EXCEL, sheet_name="Fixtures", usecols="A:H", header=0, nrows=20, dtype={"Oi": float, "Di": float})
    favourable_attacking_teams = []
    favourable_defending_teams = []

    gameweeks = fixture_df.columns[1:6]

    for col in columns:
        attack_team = fixture_sheet[f"{col}24"].value
        defend_team = fixture_sheet[f"{col}25"].value
        favourable_attacking_teams.append(
            {"Team": attack_team, 
             gameweeks[0]: fixture_df.loc[fixture_df["Team"] == attack_team, gameweeks[0]].values[0],
             gameweeks[1]: fixture_df.loc[fixture_df["Team"] == attack_team, gameweeks[1]].values[0],
             gameweeks[2]: fixture_df.loc[fixture_df["Team"] == attack_team, gameweeks[2]].values[0],
             gameweeks[3]: fixture_df.loc[fixture_df["Team"] == attack_team, gameweeks[3]].values[0],
             gameweeks[4]: fixture_df.loc[fixture_df["Team"] == attack_team, gameweeks[4]].values[0],
             "Oi": fixture_df.loc[fixture_df["Team"] == attack_team, "Oi"].values[0], 
             "Di": fixture_df.loc[fixture_df["Team"] == attack_team, "Di"].values[0],
            }
            )
        favourable_defending_teams.append(
            {"Team": defend_team, 
             gameweeks[0]: fixture_df.loc[fixture_df["Team"] == defend_team, gameweeks[0]].values[0],
             gameweeks[1]: fixture_df.loc[fixture_df["Team"] == defend_team, gameweeks[1]].values[0],
             gameweeks[2]: fixture_df.loc[fixture_df["Team"] == defend_team, gameweeks[2]].values[0],
             gameweeks[3]: fixture_df.loc[fixture_df["Team"] == defend_team, gameweeks[3]].values[0],
             gameweeks[4]: fixture_df.loc[fixture_df["Team"] == defend_team, gameweeks[4]].values[0],
             "Oi": fixture_df.loc[fixture_df["Team"] == defend_team, "Oi"].values[0], 
             "Di": fixture_df.loc[fixture_df["Team"] == defend_team, "Di"].values[0],
             }
            )

    gameweek = team_df["Games"].max()
    separator = "--------------------"
    print(f"Summary as at GW {gameweek}")
    print(separator)
    
    # team
    print(f"Average Oi: {math.floor(team_data_sheet['G23'].value * 100)/100}\n")
    print("Top 5 attacking teams")
    top_attacking_team_df = pd.DataFrame(top_attacking_team)
    print(top_attacking_team_df.sort_values(by=["Oi"], ascending=False).to_string(index=False, formatters=format_columns(top_attacking_team_df)))
    
    print(separator)
    print(f"Average Di: {math.floor(team_data_sheet['H23'].value * 100)/100}\n")
    print("Top 5 defending teams")
    top_defending_team_df = pd.DataFrame(top_defending_team)
    print(top_defending_team_df.sort_values(by=["Di"], ascending=True).to_string(index=False, formatters=format_columns(top_defending_team_df)))
    print(separator)

    # fixture
    print("Top 5 favourable attacking teams")
    print(f"Average aggregated Di: {math.floor(fixture_sheet['H23'].value * 100)/100}\n")
    favourable_attacking_teams_df = pd.DataFrame(favourable_attacking_teams)
    print(favourable_attacking_teams_df.sort_values(by=["Di"], ascending=False).to_string(index=False, formatters=format_columns(favourable_attacking_teams_df)))   
    print(separator)
    
    print("Top 5 favourable defending teams")
    print(f"Average aggregated Oi: {math.floor(fixture_sheet['G23'].value * 100)/100}\n" )
    favourable_defending_teams_df = pd.DataFrame(favourable_defending_teams)
    print(favourable_defending_teams_df.sort_values(by=["Oi"], ascending=True).to_string(index=False, formatters=format_columns(favourable_defending_teams_df)))



def update_player_last5():
    r = requests.get(BASE_URL).json()   
    players = r["elements"] # 1 to 642
    gw = get_gameweek()
    gws = range(gw-4, gw+1) if gw > 5 else range(1, gw+1)
    print([f"GW{gw}" for gw in gws])
    gks, defs, mids, fwds = [], [], [], []
    for player in players:
        if player["minutes"] > MINUTES_THRESHOLD:
            id = player["id"]
            individual = requests.get(f"https://fantasy.premierleague.com/api/element-summary/{id}/").json()["history"]
            start_round = individual[0]["round"]
            matches = len(individual)
            if matches < 5:
                gameweeks = range(0, matches-start_round+1)
            else:
                gameweeks = range(gw-start_round-4, gw-start_round+1) if gw > 5 else range(0, gw-start_round+1)
            points = dict.fromkeys(gws, 0)
            cs, goals_conceded, xg_conceded, saves, bonus, xg, g, xa, a, xgi = 0,0,0,0,0,0,0,0,0,0
            total_starts, total_games,  total_points = 0,0,0
            for i in gameweeks:
                gw_stat = individual[i]
                points[i+1] = gw_stat["total_points"]
                total_points += gw_stat["total_points"]
                total_starts += gw_stat["starts"]
                total_games += 1 if gw_stat["minutes"] > 0 else 0
                cs += gw_stat["clean_sheets"]
                goals_conceded += gw_stat["goals_conceded"]
                xg_conceded += float(gw_stat["expected_goals_conceded"])
                saves += gw_stat["saves"]
                bonus += gw_stat["bonus"]
                xg += float(gw_stat["expected_goals"])
                g += gw_stat["goals_scored"]
                xa += float(gw_stat["expected_assists"])
                a += gw_stat["assists"]
                xgi += float(gw_stat["expected_goal_involvements"])
            outfield_list = [player["web_name"],
                    teams[player["team"]],
                    (player["now_cost"] / 10.0),
                    position[player["element_type"]],
                    float(player["form"]),
                    total_starts,
                    total_games,
                    points[gws[0]],
                    points[gws[1]],
                    points[gws[2]],
                    #points[gws[3]],
                    #points[gws[4]],
                    total_points,
                    round(total_points / (player["now_cost"] / 10.0), 2),
                    g,
                    float(xg),
                    float(xg - g),
                    a,
                    float(xa),
                    float(xa - a),
                    float(xgi),
                    float(xgi - g - a),
                    bonus,]
            if player["element_type"] == 1:
                gks.append([
                    player["web_name"],
                    teams[player["team"]],
                    (player["now_cost"] / 10.0),
                    position[player["element_type"]],
                    total_starts,
                    total_games,
                    points[gws[0]],
                    points[gws[1]],
                    points[gws[2]],
                    #points[gws[3]],
                    #points[gws[4]],
                    total_points,
                    round(total_points / (player["now_cost"] / 10.0), 2),
                    int(cs),
                    goals_conceded,
                    float(xg_conceded),
                    float(xg_conceded - goals_conceded),
                    saves,
                    bonus,    
                ])
            elif player["element_type"] == 2:
                defs.append(outfield_list)
            elif player["element_type"] == 3:
                mids.append(outfield_list)
            elif player["element_type"] == 4:
                fwds.append(outfield_list)
            print("Finished updating " + player["web_name"])

    gk_column = ["Name", "Team", "Price", "Pos", "Starts", "Games", "Total_Points", "Points/$",
    "CS", "Goals Conceded", "xG Conceded", "xG Prevented", "Saves","Bonus"]

    gk_column = gk_column[0:6] + [f"GW{gw}" for gw in gws] + gk_column[6:]    

    column = [
    "Name", "Team", "Price", "Pos", "Form", "Starts", "Games", "Total_Points", "Points/$",
    "G", "xG", "xGDiff", "A", "xA", "xADiff", "xGI", "xGIDiff", "Bonus"
    ]

    column = column[0:7] + [f"GW{gw}" for gw in gws] + column[7:]

    gk_df = pd.DataFrame(gks, columns=gk_column)
    gk_df = gk_df.sort_values(by=["Points/$", "xG Prevented"], ascending=[False, False])

    def_df = pd.DataFrame(defs, columns=column)
    def_df = def_df.sort_values(by=["Points/$", "xGI"], ascending=[False, False])

    mid_df = pd.DataFrame(mids, columns=column)
    mid_df = mid_df.sort_values(by=["Points/$", "xGI"], ascending=[False, False])
    
    fwd_df = pd.DataFrame(fwds, columns=column)
    fwd_df = fwd_df.sort_values(by=["Points/$", "xGI"], ascending=[False, False])

    wb = load_workbook(LAST5_EXCEL)
    gk_sheet = wb["GK"]
    def_sheet = wb["DEF"]
    mid_sheet = wb["MID"]
    fwd_sheet = wb["FWD"]

    write_df_to_sheet(gk_df, gk_sheet)
    write_df_to_sheet(def_df, def_sheet)
    write_df_to_sheet(mid_df, mid_sheet)
    write_df_to_sheet(fwd_df, fwd_sheet)
    print("Finished updated player data.")
    wb.save(LAST5_EXCEL)


def update_results_last5():
    gw = get_gameweek()
    gws = [gw-4, gw-3, gw-2, gw-1, gw]
    fixtures = requests.get("https://fantasy.premierleague.com/api/fixtures/").json()


def update_my_team_last5():
    gw = get_gameweek()
    url = f"https://fantasy.premierleague.com/api/entry/{ID}/event/{gw}/picks/"
    team = requests.get(url).json()["picks"]
    player_ids = [player["element"] for player in team]
    r = requests.get(BASE_URL).json()["elements"]
    players = [player for player in r if player["id"] in player_ids]

    wb = load_workbook(LAST5_EXCEL)
    gk_df = pd.read_excel(LAST5_EXCEL, sheet_name="GK", usecols=LAST5_GK_COLS, header=0, nrows=GK_COUNT)
    def_df = pd.read_excel(LAST5_EXCEL, sheet_name="DEF", usecols=LAST5_DEF_COLS, header=0, nrows=DEF_COUNT)
    mid_df = pd.read_excel(LAST5_EXCEL, sheet_name="MID", usecols=LAST5_MID_COLS, header=0, nrows=MID_COUNT)
    fwd_df = pd.read_excel(LAST5_EXCEL, sheet_name="FWD", usecols=LAST5_FWD_COLS, header=0, nrows=FWD_COUNT)

    team_gk_rows = []
    team_def_rows = []
    team_mid_rows = []
    team_fwd_rows = []

    for player in players:
        if player["element_type"] == 1:
            team_gk_rows.append(gk_df.loc[gk_df["Name"] == player["web_name"]])
        elif player["element_type"] == 2:
            team_def_rows.append(def_df.loc[def_df["Name"] == player["web_name"]])
        elif player["element_type"] == 3:
            team_mid_rows.append(mid_df.loc[mid_df["Name"] == player["web_name"]])
        elif player["element_type"] == 4:
            team_fwd_rows.append(fwd_df.loc[fwd_df["Name"] == player["web_name"]])
        print("Finished getting " + player["web_name"])

    team_gk_df = pd.concat(team_gk_rows, ignore_index=True)
    team_def_df = pd.concat(team_def_rows, ignore_index=True)
    team_mid_df = pd.concat(team_mid_rows, ignore_index=True)
    team_fwd_df = pd.concat(team_fwd_rows, ignore_index=True)

    team_df = pd.concat([team_def_df, team_mid_df, team_fwd_df, team_gk_df], ignore_index=True, sort=False)
    team_df = team_df.sort_values(by=["Points/$", "xGI"], ascending=[False, False])

    team = wb["My Team"]
    write_df_to_sheet(team_df, team)
    wb.save(LAST5_EXCEL)
    print("Finished updating my team - last 5")


def update_teams_season():
    print("Start updating teams season data.")
    #web scrapping
    url = "https://fpl.page/bonus"
    page = requests.get(url)

    soup = BeautifulSoup(page.text, "html.parser")
    
    fixtures = soup.find_all("li", class_="fixture-item")
    
    teams = []

    for fixture in fixtures:
        home_team = fixture.find("span", class_="home-text").text.strip()
        away_team = fixture.find("span", class_="away-text").text.strip()
        score_box = fixture.find("span", class_="score-box")
        if score_box:
            scores = score_box.text.split("-")
            scores = list(map(int, scores))
            
            teams.append({"Team": home_team, "Side": "H", "GF": scores[0], "GA": scores[1]})
            teams.append({"Team": away_team, "Side": "A", "GF": scores[1], "GA": scores[0]})
            print(f"{home_team} {scores[0]} - {scores[1]} {away_team}")

    print("Finished getting gameweek data")
    #update excel
    wb = load_workbook(SEASON_EXCEL)
    sheet = wb["Teams"]

    df = pd.read_excel(SEASON_EXCEL, sheet_name="Teams", usecols=SEASON_TEAMS_COLS, header=0, nrows=20)
    for team in teams:
        team_name = team["Team"]
        team_row = df.index[df["Team"] == team_name].tolist()[0]
        df.loc[team_row, "GF"] += team["GF"]
        df.loc[team_row, "GA"] += team["GA"]
        df.loc[team_row, "Games"] += 1
        df.loc[team_row, "Avg.GF"] = df.loc[team_row, "GF"] / df.loc[team_row, "Games"]
        df.loc[team_row, "Avg.GA"] = df.loc[team_row, "GA"] / df.loc[team_row, "Games"]

        #update home side
        if team["Side"] == "H":
            df.loc[team_row, "H_GF"] += team["GF"]
            df.loc[team_row, "H_GA"] += team["GA"]
            df.loc[team_row, "H_Games"] += 1
        #update away side
        elif team["Side"] == "A":
            df.loc[team_row, "A_GF"] += team["GF"]
            df.loc[team_row, "A_GA"] += team["GA"]
            df.loc[team_row, "A_Games"] += 1
        print(f"Updated data of {team_name}")

    df["H_Avg.GF"] = df["H_GF"] / df["H_Games"]
    df["H_Avg.GA"] = df["H_GA"] / df["H_Games"]
    df["A_Avg.GF"] = df["A_GF"] / df["A_Games"]
    df["A_Avg.GA"] = df["A_GA"] / df["A_Games"]
    df["Avg.GF"] = df["GF"] / df["Games"]
    df["Avg.GA"] = df["GA"] / df["Games"]

    lod = df["GF"].sum() / df["Games"].sum()
    df["Oi"] = df["Avg.GF"] / lod
    df["Di"] = df["Avg.GA"] / lod

    h_lod = df["H_GF"].sum() / df["H_Games"].sum()
    df["H_Oi"] = df["H_Avg.GF"] / h_lod
    df["H_Di"] = df["H_Avg.GA"] / h_lod

    a_lod = df["A_GF"].sum() / df["A_Games"].sum()
    df["A_Oi"] = df["A_Avg.GF"] / a_lod
    df["A_Di"] = df["A_Avg.GA"] / a_lod

    cols = ["H_Avg.GF", "H_Avg.GA", "H_Oi", "H_Di", "A_Avg.GF", "A_Avg.GA", "A_Oi", "A_Di", "Avg.GF", "Avg.GA", "Oi", "Di"]
    for col in cols:
        df[col] = df[col].round(2)

    sheet["B23"] = df["H_GF"].sum()
    sheet["C23"] = df["H_GA"].sum()
    sheet["D23"] = df["A_GF"].sum()
    sheet["E23"] = df["A_GA"].sum()
    sheet["F23"] = df["GF"].sum()
    sheet["G23"] = df["GA"].sum()
    sheet["H23"] = df["H_Games"].sum()
    sheet["I23"] = df["A_Games"].sum()
    sheet["J23"] = df["Games"].sum()

    sheet["K23"] = df["H_Avg.GF"].mean()
    sheet["L23"] = df["H_Avg.GA"].mean()
    sheet["M23"] = df["A_Avg.GF"].mean()
    sheet["N23"] = df["A_Avg.GA"].mean()
    sheet["O23"] = df["Avg.GF"].mean()
    sheet["P23"] = df["Avg.GA"].mean()
    sheet["Q23"] = df["H_Oi"].mean()
    sheet["R23"] = df["H_Di"].mean()
    sheet["S23"] = df["A_Oi"].mean()
    sheet["T23"] = df["A_Di"].mean()
    sheet["U23"] = df["Oi"].mean()
    sheet["V23"] = df["Di"].mean()

    sheet["B24"] = h_lod
    sheet["E24"] = a_lod
    sheet["H24"] = lod

    home_top_attacking = df.sort_values(by=["H_Oi"], ascending=False)["Team"].head(5).tolist()
    away_top_attacking = df.sort_values(by=["A_Oi"], ascending=False)["Team"].head(5).tolist()
    top_attacking = df.sort_values(by=["Oi"], ascending=False)["Team"].head(5).tolist()

    home_top_defending = df.sort_values(by=["H_Di"], ascending=True)["Team"].head(5).tolist()
    away_top_defending = df.sort_values(by=["A_Di"], ascending=True)["Team"].head(5).tolist()
    top_defending = df.sort_values(by=["Di"], ascending=True)["Team"].head(5).tolist()


    low_oi_index = df.sort_values(by=["Oi"], ascending=True)["Oi"].head(5).index.tolist()
    high_oi_index = df.sort_values(by=["Oi"], ascending=False)["Oi"].head(5).index.tolist()
    high_di_index = df.sort_values(by=["Di"], ascending=False)["Di"].head(5).index.tolist()
    low_di_index = df.sort_values(by=["Di"], ascending=True)["Di"].head(5).index.tolist()

    columns = ["B", "C", "D", "E", "F"]

    for i, col in enumerate(columns):
        sheet[f"{col}26"] = home_top_attacking[i]
        sheet[f"{col}27"] = away_top_attacking[i]
        sheet[f"{col}28"] = top_attacking[i]

        sheet[f"{col}30"] = home_top_defending[i]
        sheet[f"{col}31"] = away_top_defending[i]
        sheet[f"{col}32"] = top_defending[i]

    for r_idx, row in enumerate(df.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    for high_oi_idx, low_oi_idx, high_di_idx, low_di_idx in zip(high_oi_index, low_oi_index, high_di_index, low_di_index):
        sheet.cell(row=low_oi_idx + 2, column=COL_U).fill = redFill
        sheet.cell(row=high_di_idx + 2, column=COL_V).fill = redFill

        sheet.cell(row=high_oi_idx + 2, column=COL_U).fill = blueFill
        sheet.cell(row=high_oi_idx + 2, column=COL_A).fill = blueFill

        sheet.cell(row=low_di_idx + 2, column=COL_V).fill = greenFill
        sheet.cell(row=low_di_idx + 2, column=COL_A).fill = greenFill

        if high_oi_idx in low_di_index:
            sheet.cell(row=high_oi_idx + 2, column=COL_A).fill = yellowFill
        
        if low_di_idx in high_oi_index:
            sheet.cell(row=low_di_idx + 2, column=COL_A).fill = yellowFill
        
        if low_oi_idx in high_di_index:
            sheet.cell(row=low_oi_idx + 2, column=COL_A).fill = redFill
        
        if high_di_idx in low_oi_index:
            sheet.cell(row=high_di_idx + 2, column=COL_A).fill = redFill

    print("Finished updating team gameweek data.")    
    wb.save(SEASON_EXCEL)


def update_player_season():
    gk_df = pd.read_excel(SEASON_EXCEL, sheet_name="GK", usecols=SEASON_GK_COLS, header=0, nrows=GK_COUNT)
    def_df = pd.read_excel(SEASON_EXCEL, sheet_name="DEF", usecols=SEASON_DEF_COLS, header=0, nrows=DEF_COUNT)
    mid_df = pd.read_excel(SEASON_EXCEL, sheet_name="MID", usecols=SEASON_MID_COLS, header=0, nrows=MID_COUNT)
    fwd_df = pd.read_excel(SEASON_EXCEL, sheet_name="FWD", usecols=SEASON_FWD_COLS, header=0, nrows=FWD_COUNT)

    r = requests.get(BASE_URL).json() 
    players = r["elements"] # 1 to 642
    gw = get_gameweek()
    for player in players:
        if player["minutes"] > MINUTES_THRESHOLD:
            id = player["id"]
            individual = requests.get(f"https://fantasy.premierleague.com/api/element-summary/{id}/").json()["history"]
            start_round = individual[0]["round"]
            # if I forgot to update after a gw, use to following line
            # gameweeks = [i for i in range(gw-n-start_round, gw-n-start_round+1)] where n is the number of gws i missed
            gameweeks = [i for i in range(gw-start_round, gw-start_round + 1)]
            if player["element_type"] == 1:
                df = gk_df
                player_row = gk_df.index[gk_df["Name"] == player["web_name"]].tolist()[0]
            elif player["element_type"] == 2:
                df = def_df
                player_row = def_df.index[def_df["Name"] == player["web_name"]].tolist()[0]
            elif player["element_type"] == 3:
                df = mid_df
                player_row = mid_df.index[mid_df["Name"] == player["web_name"]].tolist()[0]
            elif player["element_type"] == 4:
                df = fwd_df
                player_row = fwd_df.index[fwd_df["Name"] == player["web_name"]].tolist()[0]
            for i in gameweeks:
                gw_stat = individual[i]
                side = "H" if gw_stat["was_home"] else "A"
                if player["element_type"] == 1:
                    df.iloc[player_row, df.columns.get_loc(f"{side}_Points")] += gw_stat["total_points"]
                    df.iloc[player_row, df.columns.get_loc(f"{side}_CS")] += gw_stat["clean_sheets"]
                    df.iloc[player_row, df.columns.get_loc(f"{side}_Goals Conceded")] += gw_stat["goals_conceded"]
                    df.iloc[player_row, df.columns.get_loc(f"{side}_xG Conceded")] += float(gw_stat["expected_goals_conceded"])
                    df.iloc[player_row, df.columns.get_loc(f"{side}_xG Prevented")] = df.iloc[player_row, df.columns.get_loc(f"{side}_xG Conceded")] - df.iloc[player_row, df.columns.get_loc(f"{side}_Goals Conceded")]
                    df.iloc[player_row, df.columns.get_loc(f"{side}_Saves")] += gw_stat["saves"]
                    df.iloc[player_row, df.columns.get_loc(f"{side}_Bonus")] += gw_stat["bonus"]
                else:
                    df.iloc[player_row, df.columns.get_loc(f"{side}_Points")] += gw_stat["total_points"]
                    df.iloc[player_row, df.columns.get_loc(f"{side}_G")] += gw_stat["goals_scored"]
                    df.iloc[player_row, df.columns.get_loc(f"{side}_xG")] += float(gw_stat["expected_goals"])
                    df.iloc[player_row, df.columns.get_loc(f"{side}_xGDiff")] = df.iloc[player_row, df.columns.get_loc(f"{side}_xG")] - df.iloc[player_row, df.columns.get_loc(f"{side}_G")]
                    df.iloc[player_row, df.columns.get_loc(f"{side}_A")] += gw_stat["assists"]
                    df.iloc[player_row, df.columns.get_loc(f"{side}_xA")] += float(gw_stat["expected_assists"])
                    df.iloc[player_row, df.columns.get_loc(f"{side}_xADiff")] = df.iloc[player_row, df.columns.get_loc(f"{side}_xA")] - df.iloc[player_row, df.columns.get_loc(f"{side}_A")]
                    df.iloc[player_row, df.columns.get_loc(f"{side}_xGI")] += float(gw_stat["expected_goal_involvements"])
                    df.iloc[player_row, df.columns.get_loc(f"{side}_xGIDiff")] = df.iloc[player_row, df.columns.get_loc(f"{side}_xGI")] - df.iloc[player_row, df.columns.get_loc(f"{side}_G")] - df.iloc[player_row, df.columns.get_loc(f"{side}_A")]
                    df.iloc[player_row, df.columns.get_loc(f"{side}_Bonus")] += gw_stat["bonus"]
            if player["element_type"] == 1:
                df.iloc[player_row, df.columns.get_loc("CS")] = df.iloc[player_row, df.columns.get_loc("H_CS")] + df.iloc[player_row, df.columns.get_loc("A_CS")]
                df.iloc[player_row, df.columns.get_loc("Goals Conceded")] = df.iloc[player_row, df.columns.get_loc("H_Goals Conceded")] + df.iloc[player_row, df.columns.get_loc("A_Goals Conceded")]
                df.iloc[player_row, df.columns.get_loc("xG Conceded")] = df.iloc[player_row, df.columns.get_loc("H_xG Conceded")] + df.iloc[player_row, df.columns.get_loc("A_xG Conceded")]
                df.iloc[player_row, df.columns.get_loc("xG Prevented")] = df.iloc[player_row, df.columns.get_loc("H_xG Prevented")] + df.iloc[player_row, df.columns.get_loc("A_xG Prevented")]
                df.iloc[player_row, df.columns.get_loc("H_xG Prevented / $")] = round(df.iloc[player_row, df.columns.get_loc("H_xG Prevented")] / df.iloc[player_row, df.columns.get_loc("Price")], 2)
                df.iloc[player_row, df.columns.get_loc("A_xG Prevented / $")] = round(df.iloc[player_row, df.columns.get_loc("A_xG Prevented")] / df.iloc[player_row, df.columns.get_loc("Price")], 2)
                df.iloc[player_row, df.columns.get_loc("xG Prevented / $")] = df.iloc[player_row, df.columns.get_loc("H_xG Prevented / $")] + df.iloc[player_row, df.columns.get_loc("A_xG Prevented / $")]
                df.iloc[player_row, df.columns.get_loc("Saves")] = df.iloc[player_row, df.columns.get_loc("H_Saves")] + df.iloc[player_row, df.columns.get_loc("A_Saves")]
            else: 
                df.iloc[player_row, df.columns.get_loc("Form")] = float(player["form"])
                df.iloc[player_row, df.columns.get_loc("G")] = df.iloc[player_row, df.columns.get_loc("H_G")] + df.iloc[player_row, df.columns.get_loc("A_G")]
                df.iloc[player_row, df.columns.get_loc("xG")] = df.iloc[player_row, df.columns.get_loc("H_xG")] + df.iloc[player_row, df.columns.get_loc("A_xG")]
                df.iloc[player_row, df.columns.get_loc("xGDiff")] = df.iloc[player_row, df.columns.get_loc("H_xGDiff")] + df.iloc[player_row, df.columns.get_loc("A_xGDiff")]
                df.iloc[player_row, df.columns.get_loc("A")] = df.iloc[player_row, df.columns.get_loc("H_A")] + df.iloc[player_row, df.columns.get_loc("A_A")]
                df.iloc[player_row, df.columns.get_loc("xA")] = df.iloc[player_row, df.columns.get_loc("H_xA")] + df.iloc[player_row, df.columns.get_loc("A_xA")]
                df.iloc[player_row, df.columns.get_loc("xADiff")] = df.iloc[player_row, df.columns.get_loc("H_xADiff")] + df.iloc[player_row, df.columns.get_loc("A_xADiff")]
                df.iloc[player_row, df.columns.get_loc("xGI")] = df.iloc[player_row, df.columns.get_loc("H_xGI")] + df.iloc[player_row, df.columns.get_loc("A_xGI")]
                df.iloc[player_row, df.columns.get_loc("H_xGI / $")] = round(df.iloc[player_row, df.columns.get_loc("H_xGI")] / df.iloc[player_row, df.columns.get_loc("Price")], 2)
                df.iloc[player_row, df.columns.get_loc("A_xGI / $")] = round(df.iloc[player_row, df.columns.get_loc("A_xGI")] / df.iloc[player_row, df.columns.get_loc("Price")], 2)
                df.iloc[player_row, df.columns.get_loc("xGI / $")] = df.iloc[player_row, df.columns.get_loc("H_xGI / $")] + df.iloc[player_row, df.columns.get_loc("A_xGI / $")]
                df.iloc[player_row, df.columns.get_loc("xGIDiff")] = df.iloc[player_row, df.columns.get_loc("H_xGIDiff")] + df.iloc[player_row, df.columns.get_loc("A_xGIDiff")]
            
            df.iloc[player_row, df.columns.get_loc(f"{side}_Games")] += 1 if gw_stat["minutes"] > 0 else 0
            df.iloc[player_row, df.columns.get_loc("Starts")] += gw_stat["starts"]
            df.iloc[player_row, df.columns.get_loc("Games")] = df.iloc[player_row, df.columns.get_loc("H_Games")] + df.iloc[player_row, df.columns.get_loc("A_Games")]
            df.iloc[player_row, df.columns.get_loc("Points")] = df.iloc[player_row, df.columns.get_loc("H_Points")] + df.iloc[player_row, df.columns.get_loc("A_Points")]
            df.iloc[player_row, df.columns.get_loc("Price")] = player["now_cost"] / 10.0
            df.iloc[player_row, df.columns.get_loc("H_Points/$")] = round(df.iloc[player_row, df.columns.get_loc("H_Points")] / df.iloc[player_row, df.columns.get_loc("Price")], 2)
            df.iloc[player_row, df.columns.get_loc("A_Points/$")] = round(df.iloc[player_row, df.columns.get_loc("A_Points")] / df.iloc[player_row, df.columns.get_loc("Price")], 2)
            df.iloc[player_row, df.columns.get_loc("Points/$")] = round(df.iloc[player_row, df.columns.get_loc("Points")] / df.iloc[player_row, df.columns.get_loc("Price")], 2)
            df.iloc[player_row, df.columns.get_loc("Bonus")] = df.iloc[player_row, df.columns.get_loc("H_Bonus")] + df.iloc[player_row, df.columns.get_loc("A_Bonus")]
            print(f"Finished updating {player['web_name']}")
    wb = load_workbook(SEASON_EXCEL)

    gk_sheet = wb["GK"]
    def_sheet = wb["DEF"]
    mid_sheet = wb["MID"]
    fwd_sheet = wb["FWD"]

    gk_df = gk_df.sort_values(by=["Points/$", "xG Prevented"], ascending=[False, False])
    def_df = def_df.sort_values(by=["Points/$", "xGI"], ascending=[False, False])
    mid_df = mid_df.sort_values(by=["Points/$", "xGI"], ascending=[False, False])
    fwd_df = fwd_df.sort_values(by=["Points/$", "xGI"], ascending=[False, False])

    write_df_to_sheet(gk_df, gk_sheet)
    write_df_to_sheet(def_df, def_sheet)
    write_df_to_sheet(mid_df, mid_sheet)
    write_df_to_sheet(fwd_df, fwd_sheet)

    wb.save(SEASON_EXCEL)
    print("Finished updating player season data.")



def update_my_team_season():
    gw = get_gameweek()
    url = f"https://fantasy.premierleague.com/api/entry/{ID}/event/{gw}/picks/"
    team = requests.get(url).json()["picks"]
    player_ids = [player["element"] for player in team]
    r = requests.get(BASE_URL).json()["elements"]
    players = [player for player in r if player["id"] in player_ids]

    wb = load_workbook(SEASON_EXCEL)
    gk_df = pd.read_excel(SEASON_EXCEL, sheet_name="GK", usecols=SEASON_GK_COLS, header=0, nrows=GK_COUNT)
    def_df = pd.read_excel(SEASON_EXCEL, sheet_name="DEF", usecols=SEASON_DEF_COLS, header=0, nrows=DEF_COUNT)
    mid_df = pd.read_excel(SEASON_EXCEL, sheet_name="MID", usecols=SEASON_MID_COLS, header=0, nrows=MID_COUNT)
    fwd_df = pd.read_excel(SEASON_EXCEL, sheet_name="FWD", usecols=SEASON_FWD_COLS, header=0, nrows=FWD_COUNT)

    team_gk_rows = []
    team_def_rows = []
    team_mid_rows = []
    team_fwd_rows = []

    for player in players:
        if player["element_type"] == 1:
            team_gk_rows.append(gk_df.loc[gk_df["Name"] == player["web_name"]])
        elif player["element_type"] == 2:
            team_def_rows.append(def_df.loc[def_df["Name"] == player["web_name"]])
        elif player["element_type"] == 3:
            team_mid_rows.append(mid_df.loc[mid_df["Name"] == player["web_name"]])
        elif player["element_type"] == 4:
            team_fwd_rows.append(fwd_df.loc[fwd_df["Name"] == player["web_name"]])
        print("Finished getting " + player["web_name"])

    team_gk_df = pd.concat(team_gk_rows, ignore_index=True)
    team_def_df = pd.concat(team_def_rows, ignore_index=True)
    team_mid_df = pd.concat(team_mid_rows, ignore_index=True)
    team_fwd_df = pd.concat(team_fwd_rows, ignore_index=True)

    team_df = pd.concat([team_def_df, team_mid_df, team_fwd_df, team_gk_df], ignore_index=True, sort=False)
    team_df = team_df.sort_values(by=["Points/$", "xGI"], ascending=[False, False])

    team_season = wb["My Team"]
    write_df_to_sheet(team_df, team_season)
    wb.save(SEASON_EXCEL)
    print("Finished updating my team - season")
#region Main function
if __name__ == "__main__":
    update = True
    while(update):
        print("Which timeframe would you like to update/get?")
        print("1. Current GW \t 2. Next GW \t 3. Last 5 GWs \t 4. Next 5 GWs \t 5. Season")
        time_frame = int(input("Time frame: "))
        if(time_frame >= 6 or time_frame == 0):
            print("Invalid time frame.")
            continue
        print("What data would you like to update/get?")
        if(time_frame == 1):
            print("1. Player Data \t 2. Results \t 3. My Team \t 4. Summary \t 5. New Players")
            data = int(input("Choice: "))
            if (data > 5 or data == 0):
                print("Invalid choice.")
                continue
            if(data == 1):
                #update_player_current()
                break
            elif(data == 2):
                update_results_current()
            elif(data == 3):
                #update_my_team_current()
                break
            elif(data == 4):
                #show_summary_current()
                break
            elif(data == 5):
                get_players() 
        elif(time_frame == 2):
            print("1. Fixture Projection \t 2. Player Data Projection \t 3. My Team Projection \t 4. Summary")
            data = int(input("Choice: "))
            if (data > 3 or data == 0):
                print("Invalid choice.")
                continue
            if(data == 1):
                update_fixture_next()
            elif(data == 2):
                update_player_next()
            elif(data == 3):
                update_my_team_next()
            elif(data == 4):
                # show_summary_next()
                break
        elif(time_frame == 3):
            print("1. Team Data \t 2. Results \t 3. Player Data \t 4. My Team \t 5. Summary")
            data = int(input("Choice: "))
            if (data > 5 or data == 0):
                print("Invalid choice.")
                continue
            if(data == 1):
                # update_teams_last5()
                break
            elif(data == 2):
                update_results_last5()
            elif(data == 3):
                update_player_last5()
            elif(data == 4):
                update_my_team_last5()  
            elif(data == 5):
                # show_summary_last5()
                break
        elif(time_frame == 4):
            print("1. Fixture \t 2. Player Data Projection \t 3. My Team Projection \t 4. Summary")
            data = int(input("Choice: "))
            if (data > 4 or data == 0):
                print("Invalid choice.")
                continue
            if(data == 1):
                update_fixture()
            elif(data == 2):
                update_player_next5()
            elif(data == 3):
                update_my_team_next5()
            elif(data == 4):
                # show_summary_next5()   
                break
        elif(time_frame == 5):
            print("1. Team Data \t 2. Player Data \t 3. My Team \t 4. Summary")
            data = int(input("Choice: "))
            if (data > 4 or data == 0):
                print("Invalid choice.")
                continue
            if(data == 1):
                update_teams_season()
            elif(data == 2):
                update_player_season()
            elif(data == 3):
                update_my_team_season()   
            elif(data == 4):
                # show_summary()
                break

        print("Would you like to update another data?")
        update = input("Y/N: ").lower() == "y"
    print("All the best in your FPL jouney.")  
